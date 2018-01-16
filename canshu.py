from openpyxl import load_workbook
from getpath import getpath

class Excel_Zhuce(object):
    wb = load_workbook(getpath('canshu.xlsx'))
    sheet = wb.get_sheet_by_name('注册')
    user,passwd,re_passwd,email,perfact = [0],[0],[0],[0],[sheet.cell(row=2,column=6).value]
    for i in range(3,8):
        user.append(sheet.cell(row=i,column=2).value)
        passwd.append(sheet.cell(row=i,column=3).value)
        re_passwd.append(sheet.cell(row=i,column=4).value)
        email.append(sheet.cell(row=i,column=5).value)
        perfact.append(sheet.cell(row=i,column=6).value)

    sheet1 = wb.get_sheet_by_name('主页面')
    url = sheet1.cell(row=2,column=1).value
#     print(perfact)
# Excel_Zhuce()